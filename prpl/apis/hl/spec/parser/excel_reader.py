
from openpyxl import load_workbook
import os
import logging

from prpl.apis.hl.factory import ExcelObjectFactory as HLAPIObjectFactory


class ExcelReader:
    """Excel Parser for prpl HL-API.

    It reads every sheet, namely 'Change-Log', 'Objects', 'Fields', 'Events', 'Response Codes' and 'ToC',
    ands returns its contents as a list of dictionaries, which can then be converted to Python objects
    with the aid of 'prpl.apis.hl.factory.ExcelObjectFactory'.

    Example:
        # Import module.
        from prpl.apis.hl.spec.parser import ExcelReader as HLAPIExcelParser

        # Loads parser with the specification.
        parser = HLAPIExcelParser('specs/hl-api.xlsx')

        # Parse change-log.
        change_log = parser.get_change_log()

        # Parse objects and procedures.
        procedures = parser.get_procedures()

        # Parse fields.
        fields = parser.get_fields()

        # Parse return codes.
        response_codes = parser.get_response_codes()

        # Parse events.
        events = parser.get_events()

    """

    def __init__(self, spec):
        """Initializes the ExcelReader parser.

        Args:
            spec (str): Relative path to Excel specification file.

        """

        self.spec_path = '{}/{}'.format(os.getcwd(), spec)
        self.raw_change_log = None
        self.raw_procedures = None
        self.raw_fields = None
        self.raw_response_codes = None
        self.raw_events = None
        self.raw_instances = None

        self.logger = logging.getLogger('ExcelReader')

    def _parse_sheet(self, name):
        """Parses de specified Excel sheet and returns its contents as a list of dictionaries.

        Args:
            name (str): Name of the Excel sheet to be parsed.

        Returns:
            list<dict>: Array of dictionaries with cell contents. The keys are the header names.

        """

        # Open work book.
        work_book = load_workbook(self.spec_path, data_only=True, read_only=True)

        # Init return array.
        entries = []

        # Open specified sheet.
        sheet = work_book[name]

        # Read headers.
        headers = []

        # run through row iterator
        for row in sheet.iter_rows():
          
          # use first iterator to create headers list
          if headers == []:

            # iterate over each cell in row
            for r in range(len(row)):

              # append cell value to header list
              headers.append(row[r].value)
          
          # if we have the headers, let's parse the rest of the objects
          else:

            # create empty entry
            entry = {}

            # iterate over each cell in row
            for r in range(len(row)):

              # make sure we don't capture empty rows
              if row[0].value:

                # assign value to attribute on entry object named after header
                entry[headers[r]] = row[r].value

            if entry != {}:
              # append entry to entries list
              entries.append(entry)

        # Close workbook.
        work_book.close()

        # Return found procedures.
        return entries

    def get_change_log(self):
        """Parses the 'Change-Log' Excel sheet and returns a list of HL-API versions and changes.

        Returns:
            list<dict>: Array of identified change-log versions.

        """

        # Open work book.
        work_book = load_workbook(self.spec_path, data_only=True, read_only=True)

        # Init return array.
        change_log = []

        # Open sheet.
        sheet = work_book['Change-Log']

        # Start reading at B2.
        cell = ['B', '2']

        # Iterate trough each row until two empty lines are found.
        while True:
            # Read Version Header.
            header = sheet[''.join(cell)].value

            # Quit if no more versions are available.
            if header is None:
                break

            # Parse header.
            tokens = header.split(' ')
            api_version = {'Number': tokens[1], 'Date': tokens[2][1:-1], 'Changes': []}

            # Read Version Body (Change-List).
            cell[1] = str(int(cell[1]) + 1)
            while True:
                # Parse change number.
                change_number = sheet[''.join(cell)].value

                # Quit if no more changes are available.
                if change_number is None:
                    self.logger.debug('ChangeLog - Found version "{}" ({}) with {} changes.'.format(
                        api_version['Number'],
                        api_version['Date'],
                        len(api_version['Changes'])))

                    # Skip the blank line for the next version.
                    cell[1] = str(int(cell[1]) + 1)
                    break

                # Parse change.
                cell[0] = 'C'
                change_description = sheet[''.join(cell)].value
                api_version['Changes'].append((change_number, change_description))

                # Iterate to next cell.
                cell = ['B', str(int(cell[1]) + 1)]

            change_log.append(api_version)
            self.logger.debug('ChangeLog - Parsed entry {}.'.format(api_version))

        # Close work book.
        work_book.close()

        # Return change-log.
        return change_log

    def get_procedures(self):
        """Parses the 'Objects & Methods' Excel sheet and returns a list of HL-API objects and procedures.

        Returns:
            list<dict>: Array of API objects and procedures.

        """

        return self._parse_sheet('Objects & Methods')

    def get_parameters(self):
        """Parses the 'Parameters' Excel sheet and returns as list of HL-API fields.

        Returns:
            list<dict>: Array of API fields.

        """

        return self._parse_sheet('Parameters')

    def get_data_types(self):
        """Parses the 'Data Types' Excel sheet and returns as list of HL-API fields.

        Returns:
            list<dict>: Array of API fields.

        """

        return self._parse_sheet('Data Types')

    def get_response_codes(self):
        """Parses the 'Response Codes' Excel sheet and returns as list of HL-API return codes.

         Returns:
             list<dict>: Array of API return codes.

        """

        return self._parse_sheet('Response Codes')

    def get_events(self):
        """Parses the 'Events' Excel sheet and returns a list of HL-API events.

         Returns:
             list<dict>: Array of API events.

        """

        return self._parse_sheet('Events')

    def get_instances(self):
        """Parses the 'ToC' Excel sheet and returns a list of HL-API object instances.

         Returns:
             list<dict>: Array of API instances.

        """

        return self._parse_sheet('ToC')

    def _build_objects(self):
        """Builds HL-API objects.

        It converts and links raw-data read from Excel file into HL-API objects.

        """

        logger = logging.getLogger('ObjectFactory')

        # Build objects.
        logger.info('Building API objects.\n')
        factory = HLAPIObjectFactory(self.raw_procedures,
                                     self.raw_fields,
                                     self.raw_events,
                                     self.raw_instances,
                                     self.raw_response_codes,
                                     self.raw_change_log)

        return factory.get_api()
        

    def parse(self):
        """Parses the Excel HL-API specification file.

        It reads all Excel sheets and converts into list of dictionaries.

        """
        logger = logging.getLogger('ExcelReader')

        # Parse change-log.
        logger.info('ChangeLog - Parsing started.')
        self.raw_change_log = self.get_change_log()
        logger.info('ChangeLog - Parsing finished with {} versions discovered.\n'.format(len(self.raw_change_log)))

        # Parse objects.
        logger.info('Objects - Parsing started.')
        self.raw_procedures = self.get_procedures()
        logger.info('Objects - Parsing finished with {} procedures discovered.\n'.format(len(self.raw_procedures)))

        # Parse Parameters.
        logger.info('Parameters - Parsing started.')
        self.raw_parameters = self.get_parameters()
        logger.info('Parameters - Parsing finished with {} fields discovered.\n'.format(len(self.raw_parameters)))

        # Parse Data Types
        logger.info('Data Types - Parsing started.')
        self.raw_data_types = self.get_data_types()
        logger.info('Data Types - Parsing finished with {} fields discovered.\n'.format(len(self.raw_data_types)))

        # Parse response codes.
        logger.info('Response Codes - Parsing started.')
        self.raw_response_codes = self.get_response_codes()
        logger.info('Response Codes - Parsing finished with {} response codes.\n'.format(len(self.raw_response_codes)))

        # Parse events.
        logger.info('Events - Parsing started.')
        self.raw_events = self.get_events()
        logger.info('Events - Parsing finished with {} events.\n'.format(len(self.raw_events)))

        # Parse ToC (instances).
        logger.info('ToC - Parsing started.')
        self.raw_instances = self.get_instances()
        logger.info('ToC - Parsing finished with {} instances.\n'.format(len(self.raw_instances)))

        logger.info('Excel - Parsing finished.\n')

        return self._build_objects()

