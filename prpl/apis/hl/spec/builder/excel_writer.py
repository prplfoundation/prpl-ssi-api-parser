import logging
import json
import shutil
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill,\
    NamedStyle, Border, Side, Alignment
from openpyxl.styles.fills import FILL_SOLID

SAMPLE_STUB = {"Header": {"Name": "OK"}}

HEADLINESTYLE = NamedStyle(name="headline")

HEADLINESTYLE.font = Font(name='Calibri Light',
                          color='FFFFFF')

HEADLINESTYLE.fill = PatternFill(fill_type=FILL_SOLID,
                                 start_color='63339C',
                                 end_color='63339C')

bd_full = Side(style='thin', color="000000")
bd_none = Side(style=None, color="000000")
font = Font(name='Calibri Light', color='000000', bold=True)
bl_font = Font(name='Calibri Light', color='FFFFFF', bold=True)
top_fill = PatternFill(fill_type=FILL_SOLID,
                       start_color='f2f2f2',
                       end_color='f2f2f2')

bottom_left_fill = PatternFill(fill_type=FILL_SOLID,
                               start_color='5c1a8f',
                               end_color='5c1a8f')

bottom_right_fill = PatternFill(fill_type=FILL_SOLID,
                                start_color='ffffff',
                                end_color='ffffff')

BOTTOM_RIGHT_CHANGELOG_VERSION_STYLE = NamedStyle(name="changelog_version_br")

BOTTOM_RIGHT_CHANGELOG_VERSION_STYLE.font = font

BOTTOM_RIGHT_CHANGELOG_VERSION_STYLE.fill = bottom_right_fill

BOTTOM_RIGHT_CHANGELOG_VERSION_STYLE.border = Border(
    left=bd_none, top=bd_full, right=bd_full, bottom=bd_full)

BOTTOM_RIGHT_CHANGELOG_VERSION_STYLE.alignment = Alignment(
    wrap_text=True, horizontal="left", vertical="top")


TOP_LEFT_CHANGELOG_VERSION_STYLE = NamedStyle(name="changelog_version_tl")

TOP_LEFT_CHANGELOG_VERSION_STYLE.font = font

TOP_LEFT_CHANGELOG_VERSION_STYLE.fill = top_fill

TOP_LEFT_CHANGELOG_VERSION_STYLE.border = Border(
    left=bd_full, top=bd_full, right=bd_none, bottom=bd_full)

TOP_LEFT_CHANGELOG_VERSION_STYLE.alignment = Alignment(
    wrap_text=False, horizontal="left", vertical="top")


TOP_RIGHT_CHANGELOG_VERSION_STYLE = NamedStyle(name="changelog_version_tr")

TOP_RIGHT_CHANGELOG_VERSION_STYLE.font = font

TOP_RIGHT_CHANGELOG_VERSION_STYLE.fill = top_fill

TOP_RIGHT_CHANGELOG_VERSION_STYLE.border = Border(
    left=bd_none, top=bd_full, right=bd_full, bottom=bd_full)


BOTTOM_LEFT_CHANGELOG_VERSION_STYLE = NamedStyle(name="changelog_version_bl")

BOTTOM_LEFT_CHANGELOG_VERSION_STYLE.font = bl_font

BOTTOM_LEFT_CHANGELOG_VERSION_STYLE.fill = bottom_left_fill

BOTTOM_LEFT_CHANGELOG_VERSION_STYLE.border = Border(
    left=bd_full, top=bd_full, right=bd_none, bottom=bd_full)

BOTTOM_LEFT_CHANGELOG_VERSION_STYLE.alignment = Alignment(
    wrap_text=True, horizontal="left", vertical="top")


HEADER_CELL_WIDTHS = {
    "changelog": {'A': 3, 'B': 3, 'C': 150},
    "objects": {'A': 11, 'B': 60, 'C': 22, 'D': 75, 'E': 85, 'F': 85, 'G': 85, 'H': 85},
    "fields": {'A': 11, 'B': 43, 'C': 19, 'D': 39, 'E': 19, 'F': 85, 'G': 10, 'H': 10, 'I': 10, 'J': 30, 'K': 57, 'L': 57, 'M': 77},
    "events": {'A': 11, 'B': 20, 'C': 75, 'D': 15, 'E': 85, 'F': 85, 'G': 85, 'H': 85, 'I': 85, 'J': 85},
    "response_codes": {'A': 40, 'B': 85, 'C': 85, 'D': 85},
    "toc": {'A': 6, 'B': 70, 'C': 18, 'D': 60},
}


def jsonPrettyPrint(input):
    try:
        return json.dumps(json.loads(input), indent=2)
    except:
        return ""


class ExcelWriter:
    def __init__(self, api, target_folder="specs/generated/xls/"):
        self.targetFolder = target_folder
        self.api = api

        self.version = self.api.get_version()
        self.response_codes = False

    def getFileContent(self, path):
        f = open(path, "r")
        res = f.read()
        f.close()
        return res

    def makeField(self,
                  fields_sheet,
                  api_object_layer,
                  api_object_name,
                  procedure_name,
                  resource,
                  field):
        # create a simple object

        # create rights string
        rights = ""
        if field.is_input:
            rights = "{}W".format(rights)
        if field.is_output:
            rights = "{}R".format(rights)

        # add field to sheet
        fields_sheet.append([
            api_object_layer,
            api_object_name,
            procedure_name,
            field.name,
            resource,
            field.description,
            field.type,
            rights,
            field.is_required,
            field.default_value,
            field.possible_values,
            field.format,
            field.notes
        ])

    def iterateThroughFields(self,
                             wb,
                             api_object_layer,
                             api_object_name,
                             procedure_name,
                             resource,
                             fields):

        fields_sheet = wb.get_sheet_by_name("Parameters")

        # loop through procedure's fields
        for f in fields:

            # load field object
            field = fields[f]

            # create a field entry
            self.makeField(fields_sheet, api_object_layer,
                           api_object_name, procedure_name, resource, field)

    def iterateThroughObjects(self, wb):

        object_sheet = wb.get_sheet_by_name("Objects & Methods")
        fields_sheet = wb.get_sheet_by_name("Parameters")
        events_sheet = wb.get_sheet_by_name("Events")
        toc_sheet = wb.get_sheet_by_name("ToC")
        response_code_sheet = wb.get_sheet_by_name("Response Codes")

        fields = {}

        # iterate over objects
        for api_object in self.api.objects:

            # iterate over object procedure
            for procedure in api_object.procedures:

                # load response sample stub
                sample = json.loads(json.dumps(SAMPLE_STUB))

                # check if we have a sample response
                if procedure.sample_response != "-":
                    # if so add it to stub
                    sample["Body"] = procedure.sample_response

                # dump dict to string
                sample = json.dumps(sample)

                # create resource name from resource identifier
                resource = re.sub(r'\{.+?\}\s?', "",
                                  api_object.resource.replace(".", " "))

                # append new row with values
                object_sheet.append([api_object.layer,
                                    api_object.name,
                                    procedure.name,
                                    procedure.sample_request,
                                    procedure.sample_response,
                                    sample,
                                    resource,
                                    procedure.description])

                # add fields to fields sheet
                self.iterateThroughFields(wb,
                                          api_object.layer,
                                          api_object.name,
                                          procedure.name,
                                          resource,
                                          procedure.fields)

            for instance in api_object.instances:
                toc_sheet.append([api_object.layer,
                                  api_object.name,
                                  instance.name,
                                  instance.description])

            for event in api_object.events:
                resource = api_object.name.replace(".", " ")
                prefix = "{}_".format(
                    api_object.name.upper().replace(".", "_"))
                full_name = "{}{}".format(prefix, event.name)
                example = json.loads(json.dumps(SAMPLE_STUB))

                example["Header"]["Code"] = event.code
                example["Header"]["Name"] = full_name

                if event.sample != "":
                    example["Body"] = event.sample

                example = json.dumps(example)

                events_sheet.append([api_object.layer,
                                     api_object.name,
                                     resource,
                                     event.code,
                                     prefix,
                                     event.name,
                                     full_name,
                                     event.sample,
                                     example,
                                     event.description])

        for response_code in self.api.response_codes:
            response_code_sheet.append([
                response_code.raised_by,
                response_code.name,
                response_code.sample,
                response_code.description])

        # activate auto filter on all rows
        object_sheet.auto_filter.ref = "A1:C{}".format(
            object_sheet._current_row)

        fields_sheet.auto_filter.ref = "A1:C{}".format(
            fields_sheet._current_row)

        events_sheet.auto_filter.ref = "A1:F{}".format(
            events_sheet._current_row)

        toc_sheet.auto_filter.ref = "A1:D{}".format(toc_sheet._current_row)

    def makeResponseCodesSheet(self, wb, responses):
        ws = wb.create_sheet(title="Response Codes")

        ws.append(["Raised By", "Name", "Sample", "Description"])
        for l in ['A', 'B', 'C', 'D']:
            header_cell = "{}1".format(l)
            ws[header_cell].style = HEADLINESTYLE
            ws.column_dimensions[l].width =\
                HEADER_CELL_WIDTHS["response_codes"][l]

        for name, value in responses.items():
            if name != "99":
                ws.append([value["raised_by"], name, value["content"]
                           ["application/json"]["example"], value["description"]])

    def addFieldRow(self,
                    ws,
                    content,
                    layer,
                    field_name,
                    field_values,
                    required,
                    name,
                    procedure,
                    req_example,
                    res_example,
                    resource,
                    rights):

        description = field_values["description"] if "description" in\
            field_values else ""
        default_value = field_values["default_value"] if "default_value" in\
            field_values else ""
        possible_values = field_values["possible_values"] if\
            "possible_values" in field_values else ""
        field_format = field_values["format"] if "format" in\
            field_values else ""
        notes = ""
        if default_value != "" and default_value != "-":
            notes = notes + "Default value is \"{}\". ".format(default_value)
        if possible_values != "" and possible_values != "-":
            notes = notes + \
                "Possible value are \"{}\". ".format(possible_values)
        if field_format != "" and field_format != "-":
            notes = notes + "Format is {}. ".format(field_format)

        if notes == "":
            notes = "-"

        ws.append([
                    layer,
                    name,
                    procedure.split(".")[-1],
                    field_name,
                    resource,
                    description,
                    field_values["type"],
                    rights,
                    required,
                    default_value,
                    possible_values,
                    field_format,
                    notes])

    def addToFieldList(self, field_list, layer, field_name, field_values, required, name, procedure, req_example, res_example, resource, rights):
        # check if it is a simple property
        if field_values["type"] != "object":
            field_list[field_name] = [layer, field_name, field_values, required,
                                      name, procedure, req_example, res_example, resource, rights]

        # if not, deal with sub objects
        else:
            for sub_name, sub_value in field_values["properties"].items():
                s_field_name = "{}.{}".format(field_name, sub_name)
                self.addToFieldList(field_list, layer, s_field_name, sub_value, required,
                                    name, procedure, req_example, res_example, resource, rights)
                # field_list[s_field_name] = [layer, s_field_name, sub_value, required, values["tags"][0], procedure, req_example, res_example, resource, rights]

    def makeFieldsSheet(self, wb):

        # create new sheet with correct title
        ws = wb.create_sheet(title="Fields")

        # append header row
        ws.append([
                    "Layer",
                    "Object",
                    "Procedure",
                    "Field",
                    "Resource",
                    "Description",
                    "Type",
                    "Rights",
                    "Required",
                    "Default Value",
                    "Possible Value",
                    "Format",
                    "Notes"
                    ])

        # apply style to header row
        for l in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            header_cell = "{}1".format(l)
            ws[header_cell].style = HEADLINESTYLE
            ws.column_dimensions[l].width = HEADER_CELL_WIDTHS["fields"][l]

        for name in self.api_json["paths"]:
            if "." in name:
                obj_n = "{}".format(self.api_json["paths"][name]['$ref']
                                    .replace("#/paths", "")).split(".json")[0]
                file_name = "{}/{}".format(
                                            self.inputFolder,
                                            self.api_json["paths"][name]['$ref'].replace("#/paths", ""))
                content = json.loads(self.getFileContent(file_name))
                paths = content["paths"]
                layer = content["components"]["schemas"][obj_n]["layer"]

                for procedure, values in paths.items():

                    try:
                        req_example = json.loads(
                            content["paths"][procedure]["requestBody"]["content"]["application/json"]["example"])
                    except:
                        req_example = {}
                    try:
                        res_example = json.loads(
                            content["paths"][procedure]["responses"]["99"]["content"]["application/json"]["example"])
                        if "Body" in res_example.keys():
                            res_example = res_example["Body"]
                    except:
                        res_example = {}

                    object_fields = {}

                    resource = re.sub(r'\{.+?\}\s?', "",
                                      values["tags"][0].replace(".", " "))

                    field_list = {}

                    # for request body
                    if "requestBody" in values:

                        for field_name, field_values in\
                            values["requestBody"]["content"]["application/json"]["schema"]["properties"].items():
                                required = field_name in\
                                    values["requestBody"]["content"]["application/json"]["schema"]["required"] if "required" in\
                                    values["requestBody"]["content"]["application/json"]["schema"] else "optional"
                                rights = "W"

                                self.addToFieldList(field_list,
                                                    layer,
                                                    field_name,
                                                    field_values,
                                                    required,
                                                    values["tags"][0],
                                                    procedure,
                                                    req_example,
                                                    res_example,
                                                    resource,
                                                    rights)

                                # add filters
                                filter_columns[0].append(values["tags"][0])
                                filter_columns[1].append(procedure)
                                filter_columns[2].append(field_name)

                    # for response body
                    if "responses" in values:

                        for field_name, field_values in values["responses"]["99"]["content"]["application/json"]["schema"]["properties"].items():
                            # check if we already have the field
                            # as part of requestBody
                            if field_name in field_list.keys():
                                # update only the rights section
                                field_list[field_name][9] = "RW"
                            else:
                                rights = "R"
                                # create new record
                                required = field_name in\
                                    values["responses"]["99"]["content"]["application/json"]["schema"]["required"] if "required" in values["responses"]["99"]["content"]["application/json"]["schema"] else "optional"

                                self.addToFieldList(field_list,
                                                    layer,
                                                    field_name,
                                                    field_values,
                                                    required,
                                                    values["tags"][0],
                                                    procedure,
                                                    req_example,
                                                    res_example,
                                                    resource,
                                                    rights)

                                # if field_values["type"] != "object":
                                #   field_list[field_name] = [layer, field_name, field_values, required, values["tags"][0], procedure, req_example, res_example, resource, rights]
                                # else:
                                #   for sub_name, sub_value in field_values["properties"].items():
                                #     field_name = "{}.{}".format(field_name, sub_name)

                                #     field_list[field_name] = [layer, field_name, sub_value, required, values["tags"][0], procedure, req_example, res_example, resource, rights]

                            # add filters
                            filter_columns[0].append(values["tags"][0])
                            filter_columns[1].append(procedure)
                            filter_columns[2].append(field_name)

                    # append all fields
                    for f in field_list:
                        self.addFieldRow(ws, content, *field_list[f])

                    if not self.response_codes:
                        self.makeResponseCodesSheet(wb, values["responses"])
                        self.response_codes = True

        ws.auto_filter.ref = "A1:C{}".format(len(self.api_json["paths"]))
        ws.auto_filter.add_filter_column(0, filter_columns[0])
        ws.auto_filter.add_filter_column(1, filter_columns[1])
        ws.auto_filter.add_filter_column(2, filter_columns[2])

    def makeEventsSheet(self, wb):
        ws = wb.create_sheet("Events", 3)

        ws.append(["Object", "Code", "Name", "Sample", "Description"])
        for l in ['A', 'B', 'C', 'D', 'E']:
            header_cell = "{}1".format(l)
            ws[header_cell].style = HEADLINESTYLE
            ws.column_dimensions[l].width = HEADER_CELL_WIDTHS["events"][l]

        filter_columns = [[], [], []]

        for component, component_value in self.api_json["components"]["schemas"].items():
            if "." in component:
                file_name, object_name = component_value['$ref'].split(
                    "#/components/schemas/")
                content = json.loads(self.getFileContent(
                    "{}/{}".format(self.inputFolder, file_name)))
                schema = content["components"]["schemas"][object_name]

                if "events" in schema:
                    for event_name, event_value in schema["events"].items():
                        ws.append(
                                    [object_name, event_name,
                                        event_value["code"],
                                        event_value["content"]["application/json"]["example"],
                                        event_value["description"]])
                        filter_columns[0].append(object_name)
                        filter_columns[1].append(event_value["code"])
                        filter_columns[2].append(event_name)

        ws.auto_filter.ref = "A1:C{}".format(
            len(self.api_json["components"]["schemas"]))
        ws.auto_filter.add_filter_column(0, filter_columns[0])
        ws.auto_filter.add_filter_column(1, filter_columns[1])
        ws.auto_filter.add_filter_column(2, filter_columns[2])

    def makeTOCSheet(self, wb):

        ws = wb.create_sheet("ToC", 4)

        ws.append(["Layer", "Object", "Instance", "Description"])
        for l in ['A', 'B', 'C', 'D']:
            header_cell = "{}1".format(l)
            ws[header_cell].style = HEADLINESTYLE
            ws.column_dimensions[l].width = HEADER_CELL_WIDTHS["toc"][l]

        filter_columns = [[], [], [], []]

        for name in self.api_json["paths"]:
            if "." in name:
                obj_n = "{}".format(self.api_json["paths"][name]['$ref']
                                    .replace("#/paths", "")).split(".json")[0]
                file_name = "{}/{}".format(
                    self.inputFolder, self.api_json["paths"][name]['$ref']
                        .replace("#/paths", ""))
                content = json.loads(self.getFileContent(file_name))
                if "instances" in content.keys():
                    instances = content["instances"]
                    for name, values in instances.items():
                        layer = content['components']['schemas'][obj_n]["layer"]
                        filter_columns[0].append(layer)
                        filter_columns[1].append(obj_n)
                        filter_columns[2].append(name)
                        filter_columns[3].append(values["description"])
                        ws.append([layer, obj_n, name, values["description"]])

        ws.auto_filter.ref = "A1:D{}".format(len(filter_columns[0]))
        ws.auto_filter.add_filter_column(0, filter_columns[0])
        ws.auto_filter.add_filter_column(1, filter_columns[1])
        ws.auto_filter.add_filter_column(2, filter_columns[2])
        ws.auto_filter.add_filter_column(3, filter_columns[3])

    def setBorders(self, ws, range, border, fill=None):

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        rows = ws[range]

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill

    def makeChangeLogSheet(self, wb):

        # use currently active sheet
        ws = wb.active

        # set sheet title
        ws.title = "Change-Log"

        # create border settings
        border = Border(
            top=Side(border_style=None, color='FF000000'),
            left=Side(border_style=None, color='FF000000'),
            right=Side(border_style=None, color='FF000000'),
            bottom=Side(border_style=None, color='FF000000'))

        # create fill pattern
        fill = PatternFill("solid", fgColor="FFFFFF")

        # set border and fill on initial cells
        self.setBorders(ws, "A1:Z500", border, fill)

        # reset current row to beginning of worksheet
        ws._current_row = 0

        # set column widths
        for l in ['A', 'B', 'C']:
            ws.column_dimensions[l].width = HEADER_CELL_WIDTHS["changelog"][l]

        # add empty row
        ws.append([])

        # iterate over versions in api
        for version in self.api.versions:

            # store starting row for subsequent styling
            starting_row = ws._current_row + 1

            # create title string
            title = "Version {} ({})".format(version.number, version.date)

            # add title row to sheet
            ws.append(["", title, ""])

            # set style to title row cells
            ws["B{}".format(ws._current_row)
               ].style = TOP_LEFT_CHANGELOG_VERSION_STYLE
            ws["C{}".format(ws._current_row)
               ].style = TOP_RIGHT_CHANGELOG_VERSION_STYLE

            # iterate over changes in each version
            for change in version.change_list:

                # add a row for each change
                ws.append(["", change[0], change[1]])

                # apply styles on change row cells
                ws["B{}".format(ws._current_row)
                   ].style = BOTTOM_LEFT_CHANGELOG_VERSION_STYLE
                ws["C{}".format(ws._current_row)
                   ].style = BOTTOM_RIGHT_CHANGELOG_VERSION_STYLE

            # add empty row at the end
            ws.append([])

    def createSheets(self, wb):

        # create Fields sheet
        ws = wb.create_sheet("Parameters")

        # append header row
        ws.append(["Layer", "Object", "Procedure", "Field",
                  "Resource", "Description", "Type",
                   "Rights", "Required", "Default Value",
                   "Possible Value", "Format", "Notes"])

        # apply style to header row
        for l in ['A', 'B', 'C', 'D', 'E', 'F', 'G',
                  'H', 'I', 'J', 'K', 'L', 'M']:
            header_cell = "{}1".format(l)
            ws[header_cell].style = HEADLINESTYLE
            ws.column_dimensions[l].width = HEADER_CELL_WIDTHS["fields"][l]

        # create new sheet
        ws = wb.create_sheet("Objects & Methods", 1)

        # append header row
        ws.append(["Layer", "Object", "Procedure", "Arguments",
                   "Reponse", "Sample", "Resource", "Description"])

        # set header cell style
        for l in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            header_cell = "{}1".format(l)
            ws[header_cell].style = HEADLINESTYLE
            ws.column_dimensions[l].width = HEADER_CELL_WIDTHS["objects"][l]

        ws = wb.create_sheet("Events", 3)

        ws.append(["Layer", "Object", "Resource", "Code", "Prefix",
                   "Event", "Name", "Parameters", "Sample", "Description"])
        for l in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            header_cell = "{}1".format(l)
            ws[header_cell].style = HEADLINESTYLE
            ws.column_dimensions[l].width = HEADER_CELL_WIDTHS["events"][l]

        ws = wb.create_sheet("ToC", 4)

        ws.append(["Layer", "Object", "Instance", "Description"])
        for l in ['A', 'B', 'C', 'D']:
            header_cell = "{}1".format(l)
            ws[header_cell].style = HEADLINESTYLE
            ws.column_dimensions[l].width = HEADER_CELL_WIDTHS["toc"][l]

        ws = wb.create_sheet("Response Codes", 5)

        ws.append(["Raised By", "Name", "Sample", "Description"])
        for l in ['A', 'B', 'C', 'D']:
            header_cell = "{}1".format(l)
            ws[header_cell].style = HEADLINESTYLE
            ws.column_dimensions[l].width =\
                HEADER_CELL_WIDTHS["response_codes"][l]

    def build(self):

        # make workbook

        wb = Workbook()

        self.makeChangeLogSheet(wb)

        self.createSheets(wb)

        self.iterateThroughObjects(wb)

        # self.makeFieldsSheet(wb)

        # self.makeEventsSheet(wb)

        # self.makeTOCSheet(wb)

        number = self.api.get_version()

        wb.save("{}/Prpl-SSI-API_v{}.xlsx".format(self.targetFolder, number))
